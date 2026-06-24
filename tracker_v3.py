import requests
import json
import csv
import os
import sqlite3
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

URL = 'https://www.stbet.com/stbetrest/services/online/meeting/kiron4EventResults'
WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))

DB_PATH = os.path.join(WORKSPACE_DIR, 'stbet_results.db')
SUMMARY_CSV = os.path.join(WORKSPACE_DIR, 'cricket_superover_summary.csv')
DETAILS_CSV = os.path.join(WORKSPACE_DIR, 'cricket_superover_details.csv')
EXCEL_PATH = os.path.join(WORKSPACE_DIR, 'cricket_superover_results.xlsx')

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'application/json, text/plain, */*',
    'Referer': 'https://www.stbet.com/',
    'Origin': 'https://www.stbet.com'
}

SUMMARY_FIELDNAMES = [
    'date', 'time', 'match_name', 'winner', 
    'first_ball_1st_innings', 'first_ball_2nd_innings',
    'total_runs_ou', 'total_4s_ou', 'total_6s_ou', 
    'most_4s', 'most_6s', 'wickets_lost_ou', 'meeting_code'
]

DETAILS_FIELDNAMES = [
    'date', 'time', 'match_name', 'meeting_code', 
    'event_id', 'event_name', 'selection_number', 
    'selection_name', 'result_position', 'win_amount', 'last_odd'
]

def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS meetings (
            meeting_code TEXT PRIMARY KEY,
            meeting_name TEXT,
            meeting_date TEXT,
            meeting_time TEXT,
            raw_json TEXT
        )
    ''')
    conn.commit()
    conn.close()

def get_existing_db_codes():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT meeting_code FROM meetings')
    codes = {row[0] for row in cursor.fetchall()}
    conn.close()
    return codes

def get_all_meetings_from_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT raw_json FROM meetings')
    rows = cursor.fetchall()
    conn.close()
    meetings = [json.loads(row[0]) for row in rows]
    meetings.sort(key=lambda x: (x.get('meetingDate', ''), x.get('meetingTime', '')))
    return meetings

def get_existing_csv_codes(csv_path):
    existing_codes = set()
    if os.path.exists(csv_path):
        try:
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    code = row.get('meeting_code')
                    if code:
                        existing_codes.add(code)
        except Exception:
            pass
    return existing_codes

def save_meeting_to_db(meeting):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    try:
        cursor.execute(
            'INSERT OR IGNORE INTO meetings (meeting_code, meeting_name, meeting_date, meeting_time, raw_json) VALUES (?, ?, ?, ?, ?)',
            (
                meeting.get('meetingCode'),
                meeting.get('meetingName'),
                meeting.get('meetingDate'),
                meeting.get('meetingTime'),
                json.dumps(meeting)
            )
        )
        conn.commit()
    except Exception as e:
        print(f"Database error: {e}")
    finally:
        conn.close()

def parse_summary_row(meeting):
    meeting_name = meeting.get('meetingName', 'Unknown Match')
    meeting_date = meeting.get('meetingDate', '')
    meeting_time = meeting.get('meetingTime', '')
    meeting_code = meeting.get('meetingCode', '')
    
    match_info = {
        'date': meeting_date,
        'time': meeting_time,
        'match_name': meeting_name,
        'winner': 'N/A',
        'first_ball_1st_innings': 'N/A',
        'first_ball_2nd_innings': 'N/A',
        'total_runs_ou': 'N/A',
        'total_4s_ou': 'N/A',
        'total_6s_ou': 'N/A',
        'most_4s': 'N/A',
        'most_6s': 'N/A',
        'wickets_lost_ou': 'N/A',
        'meeting_code': meeting_code
    }
    
    for event in meeting.get('events', []):
        event_name = event.get('eventName', '')
        event_positions = event.get('eventPositions', [])
        
        winning_selections = [pos.get('selectionName', '') for pos in event_positions if pos.get('resultPosition') == 1]
        winner_str = ", ".join(winning_selections) if winning_selections else 'N/A'
        
        if not winner_str or winner_str == 'N/A':
            continue
            
        if event_name == 'Match Result':
            match_info['winner'] = winner_str
        elif event_name == '1st Innings 1st Ball Outcome':
            match_info['first_ball_1st_innings'] = winner_str
        elif event_name == '2nd Innings 1st Ball Outcome':
            match_info['first_ball_2nd_innings'] = winner_str
        elif 'Total Match Runs Over/Under' in event_name:
            match_info['total_runs_ou'] = f"{event_name.split(' ')[-1]} - {winner_str}"
        elif 'Total Match 4s Over/Under' in event_name:
            match_info['total_4s_ou'] = f"{event_name.split(' ')[-1]} - {winner_str}"
        elif 'Total Match 6s Over/Under' in event_name:
            match_info['total_6s_ou'] = f"{event_name.split(' ')[-1]} - {winner_str}"
        elif 'Innings most 4s' in event_name:
            match_info['most_4s'] = winner_str
        elif 'Innings most 6s' in event_name:
            match_info['most_6s'] = winner_str
        elif 'Match wickets lost Over/Under' in event_name:
            match_info['wickets_lost_ou'] = f"{event_name.split(' ')[-1]} - {winner_str}"
            
    return match_info

def parse_detail_rows(meeting):
    meeting_name = meeting.get('meetingName', 'Unknown Match')
    meeting_date = meeting.get('meetingDate', '')
    meeting_time = meeting.get('meetingTime', '')
    meeting_code = meeting.get('meetingCode', '')
    
    rows = []
    for event in meeting.get('events', []):
        event_id = event.get('eventId', '')
        event_name = event.get('eventName', '')
        
        for pos in event.get('eventPositions', []):
            rows.append({
                'date': meeting_date,
                'time': meeting_time,
                'match_name': meeting_name,
                'meeting_code': meeting_code,
                'event_id': event_id,
                'event_name': event_name,
                'selection_number': pos.get('selectionNumber', ''),
                'selection_name': pos.get('selectionName', ''),
                'result_position': pos.get('resultPosition', ''),
                'win_amount': pos.get('winAmount', ''),
                'last_odd': pos.get('lastOdd', '')
            })
    return rows

def write_initial_headers():
    # Write summary CSV
    if not os.path.exists(SUMMARY_CSV):
        with open(SUMMARY_CSV, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=SUMMARY_FIELDNAMES)
            writer.writeheader()
    else:
        # Check if existing CSV has the old header format, if so, recreate it to match new fields
        try:
            with open(SUMMARY_CSV, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                first_row = next(reader)
            if 'first_ball_1st_innings' not in first_row:
                print("Updating summary CSV headers for 1st ball columns...")
                os.remove(SUMMARY_CSV)
                with open(SUMMARY_CSV, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=SUMMARY_FIELDNAMES)
                    writer.writeheader()
        except Exception:
            pass
            
    if not os.path.exists(DETAILS_CSV):
        with open(DETAILS_CSV, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=DETAILS_FIELDNAMES)
            writer.writeheader()

def update_excel_file():
    meetings = get_all_meetings_from_db()
    if not meetings:
        return
        
    wb = Workbook()
    
    # 1. Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    
    sum_headers = [
        'Date', 'Time', 'Match Name', 'Winner', 
        '1st Innings 1st Ball', '2nd Innings 1st Ball',
        'Total Runs (O/U)', 'Total 4s (O/U)', 'Total 6s (O/U)', 
        'Most 4s', 'Most 6s', 'Wickets Lost (O/U)', 'Meeting Code'
    ]
    ws_sum.append(sum_headers)
    
    # Styles
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    bold_font = Font(bold=True)
    
    # Apply bold header
    for col in range(1, len(sum_headers) + 1):
        ws_sum.cell(row=1, column=col).font = bold_font
        
    for meeting in meetings:
        summary_row = parse_summary_row(meeting)
        row_values = [
            summary_row['date'], summary_row['time'], summary_row['match_name'], summary_row['winner'],
            summary_row['first_ball_1st_innings'], summary_row['first_ball_2nd_innings'],
            summary_row['total_runs_ou'], summary_row['total_4s_ou'], summary_row['total_6s_ou'],
            summary_row['most_4s'], summary_row['most_6s'], summary_row['wickets_lost_ou'], summary_row['meeting_code']
        ]
        ws_sum.append(row_values)
        
        # Highlight the 1st Ball Outcomes columns in soft yellow (columns 5 and 6)
        r_num = ws_sum.max_row
        ws_sum.cell(row=r_num, column=5).fill = yellow_fill
        ws_sum.cell(row=r_num, column=6).fill = yellow_fill

    # 2. Details sheet
    ws_det = wb.create_sheet(title="Details")
    det_headers = [
        'Date', 'Time', 'Match Name', 'Meeting Code', 
        'Event ID', 'Event Name', 'Selection Number', 
        'Selection Name', 'Result Position', 'Win Amount', 'Last Odd'
    ]
    ws_det.append(det_headers)
    
    # Apply bold header
    for col in range(1, len(det_headers) + 1):
        ws_det.cell(row=1, column=col).font = bold_font
        
    for meeting in meetings:
        detail_rows = parse_detail_rows(meeting)
        for row in detail_rows:
            row_values = [
                row['date'], row['time'], row['match_name'], row['meeting_code'],
                row['event_id'], row['event_name'], row['selection_number'],
                row['selection_name'], row['result_position'], row['win_amount'], row['last_odd']
            ]
            ws_det.append(row_values)
            
            # Highlight row in soft green if it's the winning outcome (result_position == 1) 
            # of the "1st Ball Outcome" markets
            if row['event_name'] in ['1st Innings 1st Ball Outcome', '2nd Innings 1st Ball Outcome'] and row['result_position'] == 1:
                r_num = ws_det.max_row
                for col_idx in range(1, len(det_headers) + 1):
                    ws_det.cell(row=r_num, column=col_idx).fill = green_fill
                    
    # Auto-adjust column widths for readability
    for ws in [ws_sum, ws_det]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    try:
        wb.save(EXCEL_PATH)
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Saved highlighted Excel workbook to {EXCEL_PATH}")
    except PermissionError:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] WARNING: Permission denied writing to Excel workbook ({EXCEL_PATH}). Is it open in Excel? It will be updated on the next check once closed.")

def process_new_meetings(meetings):
    if not meetings:
        return False
        
    # 1. Save all incoming meetings to DB immediately
    db_codes = get_existing_db_codes()
    new_for_db = []
    for m in meetings:
        m_code = m.get('meetingCode')
        if m_code and m_code not in db_codes:
            save_meeting_to_db(m)
            new_for_db.append(m)
            
    if new_for_db:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Saved {len(new_for_db)} new matches to SQLite database.")
        
    # 2. Try to write missing matches to CSVs (only write if file is not locked)
    csv_updated = False
    try:
        all_meetings = get_all_meetings_from_db()
        
        # Check and write to SUMMARY_CSV
        existing_summary_codes = get_existing_csv_codes(SUMMARY_CSV)
        summary_to_write = [m for m in all_meetings if m.get('meetingCode') not in existing_summary_codes]
        
        if summary_to_write:
            with open(SUMMARY_CSV, 'a', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=SUMMARY_FIELDNAMES)
                for m in summary_to_write:
                    row = parse_summary_row(m)
                    writer.writerow(row)
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Appended {len(summary_to_write)} matches to summary CSV.")
            csv_updated = True
            
        # Check and write to DETAILS_CSV
        existing_details_codes = get_existing_csv_codes(DETAILS_CSV)
        details_to_write = [m for m in all_meetings if m.get('meetingCode') not in existing_details_codes]
        
        if details_to_write:
            with open(DETAILS_CSV, 'a', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=DETAILS_FIELDNAMES)
                for m in details_to_write:
                    rows = parse_detail_rows(m)
                    for r in rows:
                        writer.writerow(r)
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Appended details for {len(details_to_write)} matches to details CSV.")
            csv_updated = True
            
    except PermissionError:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] WARNING: Permission denied writing to CSV files. Are they open in Excel? They will be written on the next check once closed.")
        
    return csv_updated or len(new_for_db) > 0

def import_previous_results():
    temp_json_path = os.path.join(os.path.dirname(WORKSPACE_DIR), '.gemini', 'antigravity-cli', 'brain', '9840bf0d-c4d0-4adb-8ca5-fb86ef429712', 'scratch', 'api_response.json')
    if not os.path.exists(temp_json_path):
        temp_json_path = os.path.join(WORKSPACE_DIR, 'api_response.json')
        
    if os.path.exists(temp_json_path):
        try:
            with open(temp_json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                print(f"Found existing data file with {len(data)} records. Importing...")
                process_new_meetings(data)
        except Exception as e:
            print(f"Failed to import previous results: {e}")
            
    # Always generate Excel sheet on startup from DB data
    update_excel_file()

def main():
    print("==========================================================")
    print("STBet Cricket-SuperOver Continuous Tracker (V3 - Excel Highlight)")
    print(f"SQLite DB Location: {DB_PATH}")
    print(f"Excel Sheet (.xlsx): {EXCEL_PATH}")
    print(f"Summary CSV: {SUMMARY_CSV}")
    print(f"Details CSV: {DETAILS_CSV}")
    print("Checking API every 4 minutes...")
    print("==========================================================")
    
    init_db()
    write_initial_headers()
    import_previous_results()
    
    while True:
        try:
            response = requests.get(URL, headers=HEADERS, timeout=10)
            if response.status_code in [200, 202]:
                has_updates = process_new_meetings(response.json())
                if has_updates:
                    update_excel_file()
            else:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] API error: status {response.status_code}")
        except Exception as e:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Request error: {e}")
            
        time.sleep(240)

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\nTracker stopped by user.")
