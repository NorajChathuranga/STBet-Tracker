import requests
import json
import csv
import os
import shutil
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

URL = 'https://www.stbet.com/stbetrest/services/online/meeting/kiron4EventResults'
WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(WORKSPACE_DIR, 'data')

# Ensure data directory exists
os.makedirs(DATA_DIR, exist_ok=True)

# Data paths inside the data/ folder
MEETINGS_JSON = os.path.join(DATA_DIR, 'meetings.json')
SUMMARY_JSON = os.path.join(DATA_DIR, 'summary.json')
DETAILS_JSON = os.path.join(DATA_DIR, 'details.json')
SUMMARY_CSV = os.path.join(DATA_DIR, 'cricket_superover_summary.csv')
DETAILS_CSV = os.path.join(DATA_DIR, 'cricket_superover_details.csv')
EXCEL_PATH = os.path.join(DATA_DIR, 'cricket_superover_results.xlsx')

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'application/json, text/plain, */*',
    'Referer': 'https://www.stbet.com/',
    'Origin': 'https://www.stbet.com'
}

SUMMARY_FIELDNAMES = [
    'date', 'time', 'match_name', 'winner', 
    'first_ball_1st_innings', 'first_ball_2nd_innings',
    'total_runs_ou', 'total_4s_ou', 'total_6s_ou', 
    'most_4s', 'most_6s', 'wickets_lost_ou', 'meeting_code'
]

DETAILS_FIELDNAMES = [
    'date', 'time', 'match_name', 'meeting_code', 
    'event_id', 'event_name', 'selection_number', 
    'selection_name', 'result_position', 'win_amount', 'last_odd'
]

def migrate_root_files():
    """Copy existing files from root to data directory if not already copied."""
    for filename in ['cricket_superover_summary.csv', 'cricket_superover_details.csv', 'cricket_superover_results.xlsx']:
        root_file = os.path.join(WORKSPACE_DIR, filename)
        dest_file = os.path.join(DATA_DIR, filename)
        if os.path.exists(root_file) and not os.path.exists(dest_file):
            try:
                shutil.copy2(root_file, dest_file)
                print(f"Copied legacy file {filename} to data/ folder.")
            except Exception as e:
                print(f"Error copying {filename}: {e}")

def load_meetings():
    if os.path.exists(MEETINGS_JSON):
        try:
            with open(MEETINGS_JSON, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading meetings JSON: {e}")
    return {}

def save_meetings(meetings_dict):
    with open(MEETINGS_JSON, 'w', encoding='utf-8') as f:
        json.dump(meetings_dict, f, indent=2)

def parse_summary_row(meeting):
    meeting_name = meeting.get('meetingName', 'Unknown Match')
    meeting_date = meeting.get('meetingDate', '')
    meeting_time = meeting.get('meetingTime', '')
    meeting_code = meeting.get('meetingCode', '')
    
    match_info = {
        'date': meeting_date,
        'time': meeting_time,
        'match_name': meeting_name,
        'winner': 'N/A',
        'first_ball_1st_innings': 'N/A',
        'first_ball_2nd_innings': 'N/A',
        'total_runs_ou': 'N/A',
        'total_4s_ou': 'N/A',
        'total_6s_ou': 'N/A',
        'most_4s': 'N/A',
        'most_6s': 'N/A',
        'wickets_lost_ou': 'N/A',
        'meeting_code': meeting_code
    }
    
    for event in meeting.get('events', []):
        event_name = event.get('eventName', '')
        event_positions = event.get('eventPositions', [])
        
        winning_selections = [pos.get('selectionName', '') for pos in event_positions if pos.get('resultPosition') == 1]
        winner_str = ", ".join(winning_selections) if winning_selections else 'N/A'
        
        if not winner_str or winner_str == 'N/A':
            continue
            
        if event_name == 'Match Result':
            match_info['winner'] = winner_str
        elif event_name == '1st Innings 1st Ball Outcome':
            match_info['first_ball_1st_innings'] = winner_str
        elif event_name == '2nd Innings 1st Ball Outcome':
            match_info['first_ball_2nd_innings'] = winner_str
        elif 'Total Match Runs Over/Under' in event_name:
            match_info['total_runs_ou'] = f"{event_name.split(' ')[-1]} - {winner_str}"
        elif 'Total Match 4s Over/Under' in event_name:
            match_info['total_4s_ou'] = f"{event_name.split(' ')[-1]} - {winner_str}"
        elif 'Total Match 6s Over/Under' in event_name:
            match_info['total_6s_ou'] = f"{event_name.split(' ')[-1]} - {winner_str}"
        elif 'Innings most 4s' in event_name:
            match_info['most_4s'] = winner_str
        elif 'Innings most 6s' in event_name:
            match_info['most_6s'] = winner_str
        elif 'Match wickets lost Over/Under' in event_name:
            match_info['wickets_lost_ou'] = f"{event_name.split(' ')[-1]} - {winner_str}"
            
    return match_info

def parse_detail_rows(meeting):
    meeting_name = meeting.get('meetingName', 'Unknown Match')
    meeting_date = meeting.get('meetingDate', '')
    meeting_time = meeting.get('meetingTime', '')
    meeting_code = meeting.get('meetingCode', '')
    
    rows = []
    for event in meeting.get('events', []):
        event_id = event.get('eventId', '')
        event_name = event.get('eventName', '')
        
        for pos in event.get('eventPositions', []):
            rows.append({
                'date': meeting_date,
                'time': meeting_time,
                'match_name': meeting_name,
                'meeting_code': meeting_code,
                'event_id': event_id,
                'event_name': event_name,
                'selection_number': pos.get('selectionNumber', ''),
                'selection_name': pos.get('selectionName', ''),
                'result_position': pos.get('resultPosition', ''),
                'win_amount': pos.get('winAmount', ''),
                'last_odd': pos.get('lastOdd', '')
            })
    return rows

def write_initial_headers():
    if not os.path.exists(SUMMARY_CSV):
        with open(SUMMARY_CSV, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=SUMMARY_FIELDNAMES)
            writer.writeheader()
    if not os.path.exists(DETAILS_CSV):
        with open(DETAILS_CSV, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=DETAILS_FIELDNAMES)
            writer.writeheader()

def update_parsed_jsons(meetings_dict):
    """Generates summary.json and details.json from the full meetings dictionary."""
    meetings_list = list(meetings_dict.values())
    # Sort by date and time
    meetings_list.sort(key=lambda x: (x.get('meetingDate', ''), x.get('meetingTime', '')))
    
    summaries = []
    details = []
    
    for m in meetings_list:
        summaries.append(parse_summary_row(m))
        details.extend(parse_detail_rows(m))
        
    with open(SUMMARY_JSON, 'w', encoding='utf-8') as f:
        json.dump(summaries, f, indent=2)
        
    with open(DETAILS_JSON, 'w', encoding='utf-8') as f:
        json.dump(details, f, indent=2)
        
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Saved updated summary.json ({len(summaries)} rows) and details.json ({len(details)} rows)")

def update_excel_file(meetings_dict):
    meetings_list = list(meetings_dict.values())
    if not meetings_list:
        return
        
    meetings_list.sort(key=lambda x: (x.get('meetingDate', ''), x.get('meetingTime', '')))
    
    wb = Workbook()
    
    # 1. Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    
    sum_headers = [
        'Date', 'Time', 'Match Name', 'Winner', 
        '1st Innings 1st Ball', '2nd Innings 1st Ball',
        'Total Runs (O/U)', 'Total 4s (O/U)', 'Total 6s (O/U)', 
        'Most 4s', 'Most 6s', 'Wickets Lost (O/U)', 'Meeting Code'
    ]
    ws_sum.append(sum_headers)
    
    # Styles
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    bold_font = Font(bold=True)
    
    for col in range(1, len(sum_headers) + 1):
        ws_sum.cell(row=1, column=col).font = bold_font
        
    for meeting in meetings_list:
        summary_row = parse_summary_row(meeting)
        row_values = [
            summary_row['date'], summary_row['time'], summary_row['match_name'], summary_row['winner'],
            summary_row['first_ball_1st_innings'], summary_row['first_ball_2nd_innings'],
            summary_row['total_runs_ou'], summary_row['total_4s_ou'], summary_row['total_6s_ou'],
            summary_row['most_4s'], summary_row['most_6s'], summary_row['wickets_lost_ou'], summary_row['meeting_code']
        ]
        ws_sum.append(row_values)
        
        r_num = ws_sum.max_row
        ws_sum.cell(row=r_num, column=5).fill = yellow_fill
        ws_sum.cell(row=r_num, column=6).fill = yellow_fill

    # 2. Details sheet
    ws_det = wb.create_sheet(title="Details")
    det_headers = [
        'Date', 'Time', 'Match Name', 'Meeting Code', 
        'Event ID', 'Event Name', 'Selection Number', 
        'Selection Name', 'Result Position', 'Win Amount', 'Last Odd'
    ]
    ws_det.append(det_headers)
    
    for col in range(1, len(det_headers) + 1):
        ws_det.cell(row=1, column=col).font = bold_font
        
    for meeting in meetings_list:
        detail_rows = parse_detail_rows(meeting)
        for row in detail_rows:
            row_values = [
                row['date'], row['time'], row['match_name'], row['meeting_code'],
                row['event_id'], row['event_name'], row['selection_number'],
                row['selection_name'], row['result_position'], row['win_amount'], row['last_odd']
            ]
            ws_det.append(row_values)
            
            if row['event_name'] in ['1st Innings 1st Ball Outcome', '2nd Innings 1st Ball Outcome'] and row['result_position'] == 1:
                r_num = ws_det.max_row
                for col_idx in range(1, len(det_headers) + 1):
                    ws_det.cell(row=r_num, column=col_idx).fill = green_fill
                    
    for ws in [ws_sum, ws_det]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    try:
        wb.save(EXCEL_PATH)
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Saved Excel workbook to {EXCEL_PATH}")
    except PermissionError:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] WARNING: Permission denied writing to Excel workbook. File might be open.")

def append_to_csvs(new_meetings):
    if not new_meetings:
        return
        
    # Sort new meetings to append in chronological order
    new_meetings.sort(key=lambda x: (x.get('meetingDate', ''), x.get('meetingTime', '')))
    
    try:
        with open(SUMMARY_CSV, 'a', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=SUMMARY_FIELDNAMES)
            for m in new_meetings:
                writer.writerow(parse_summary_row(m))
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Appended {len(new_meetings)} matches to summary CSV.")
        
        with open(DETAILS_CSV, 'a', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=DETAILS_FIELDNAMES)
            for m in new_meetings:
                rows = parse_detail_rows(m)
                for r in rows:
                    writer.writerow(r)
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Appended details for {len(new_meetings)} matches to details CSV.")
    except PermissionError:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] WARNING: Permission denied writing to CSV files.")

def load_env_file():
    """Manually parse .env file if it exists to avoid external dependencies like python-dotenv."""
    env_path = os.path.join(WORKSPACE_DIR, '.env')
    if os.path.exists(env_path):
        try:
            with open(env_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        key, val = line.split('=', 1)
                        key = key.strip()
                        val = val.strip().strip("'").strip('"')
                        os.environ[key] = val
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Loaded environment variables from .env file.")
        except Exception as e:
            print(f"Error loading .env file: {e}")

def fetch_url_data(url):
    """Fetches JSON data from the target URL, trying configured proxy providers sequentially until one succeeds."""
    proxy_url = os.environ.get('PROXY_URL')
    scraper_api_key = os.environ.get('SCRAPER_API_KEY')
    zenrows_api_key = os.environ.get('ZENROWS_API_KEY')
    
    attempts = []
    
    # 1. ScraperAPI (https://www.scraperapi.com)
    if scraper_api_key:
        scraper_url = f"http://api.scraperapi.com?api_key={scraper_api_key}&url={url}"
        attempts.append(('ScraperAPI', lambda: requests.get(scraper_url, timeout=30)))
        
    # 2. ZenRows (https://www.zenrows.com)
    if zenrows_api_key:
        zenrows_url = "https://api.zenrows.com/v1/"
        params = {
            'apikey': zenrows_api_key,
            'url': url,
            'premium_proxy': 'true',
            'proxy_country': 'us'
        }
        attempts.append(('ZenRows', lambda: requests.get(zenrows_url, params=params, timeout=30)))
        
    # 3. Standard HTTP/HTTPS Proxy (e.g. Webshare rotating proxy)
    if proxy_url:
        clean_display_url = proxy_url.split('@')[-1] if '@' in proxy_url else proxy_url
        proxies = {
            "http": proxy_url,
            "https": proxy_url
        }
        attempts.append((f'Proxy ({clean_display_url})', lambda: requests.get(url, headers=HEADERS, proxies=proxies, timeout=20)))

    # 4. Fallback to local proxies.txt file
    proxies_txt_path = os.path.join(WORKSPACE_DIR, 'proxies.txt')
    if os.path.exists(proxies_txt_path):
        try:
            with open(proxies_txt_path, 'r', encoding='utf-8') as f:
                lines = [line.strip() for line in f if line.strip() and not line.strip().startswith('#')]
            if lines:
                import random
                selected_proxy = random.choice(lines)
                clean_display = selected_proxy.split('@')[-1] if '@' in selected_proxy else selected_proxy
                proxies = {
                    "http": selected_proxy,
                    "https": selected_proxy
                }
                attempts.append((f'proxies.txt ({clean_display})', lambda: requests.get(url, headers=HEADERS, proxies=proxies, timeout=15)))
        except Exception as e:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Error reading proxies.txt: {e}")
            
    # 5. Direct connection (default fallback)
    attempts.append(('Direct Connection', lambda: requests.get(url, headers=HEADERS, timeout=15)))
    
    last_response = None
    last_error = None
    
    for name, fetch_action in attempts:
        try:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Fetching from API via {name}...")
            res = fetch_action()
            last_response = res
            if res.status_code in [200, 202]:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] Success via {name}!")
                return res
            else:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] {name} returned status {res.status_code}")
                last_error = f"HTTP {res.status_code}"
        except Exception as e:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] {name} failed: {e}")
            last_error = str(e)
            
    if last_response is not None:
        return last_response
        
    raise Exception(f"All fetch attempts failed. Last error: {last_error}")

def check_and_update():
    try:
        response = fetch_url_data(URL)
        if response.status_code not in [200, 202]:
            print(f"API Error: HTTP status {response.status_code}")
            try:
                print(f"Response preview: {response.text[:200]}")
            except Exception:
                pass
            return False
            
        data = response.json()
        if not isinstance(data, list):
            print("API Error: Expected a list of meetings, got something else.")
            try:
                print(f"Response preview: {response.text[:200]}")
            except Exception:
                pass
            return False
            
        meetings_dict = load_meetings()
        new_meetings = []
        
        for meeting in data:
            m_code = meeting.get('meetingCode')
            if m_code and m_code not in meetings_dict:
                meetings_dict[m_code] = meeting
                new_meetings.append(meeting)
                
        if new_meetings:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Found {len(new_meetings)} new match results.")
            save_meetings(meetings_dict)
            update_parsed_jsons(meetings_dict)
            update_excel_file(meetings_dict)
            append_to_csvs(new_meetings)
            return True
        else:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] No new match results found.")
            return False
    except Exception as e:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Error checking/updating: {e}")
        return False

def is_pid_running(pid):
    """Check if a process with the given PID is currently running (cross-platform)."""
    if pid <= 0:
        return False
    if os.name == 'nt':
        import subprocess
        try:
            out = subprocess.run(['tasklist', '/FI', f'PID eq {pid}'], capture_output=True, text=True)
            return str(pid) in out.stdout
        except Exception:
            return True # Fallback to True to be safe
    else:
        import errno
        try:
            os.kill(pid, 0)
        except OSError as err:
            if err.errno == errno.ESRCH:
                return False
            return True # EPERM or other errors mean process is running but not signalable
        return True

def git_commit_and_push():
    """Automatically commit and push changes in data/ folder to GitHub if running locally or on a VM."""
    import subprocess
    
    # 1. Check if git index.lock exists (indicates Git is already working)
    git_lock_path = os.path.join(WORKSPACE_DIR, '.git', 'index.lock')
    if os.path.exists(git_lock_path):
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Git is currently busy (.git/index.lock exists). Skipping commit...")
        return

    # 2. Check and handle our custom commit lock file to prevent concurrent executions
    commit_lock_path = os.path.join(WORKSPACE_DIR, 'git_commit.lock')
    if os.path.exists(commit_lock_path):
        try:
            with open(commit_lock_path, 'r') as f:
                pid = int(f.read().strip())
            if is_pid_running(pid):
                print(f"[{datetime.now().strftime('%H:%M:%S')}] Another git commit/push process (PID {pid}) is already running. Skipping...")
                return
            else:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] Stale commit lock file found (PID {pid} not running). Removing lock and proceeding...")
                os.remove(commit_lock_path)
        except Exception:
            # If lock file is empty or corrupted, clean it
            try:
                os.remove(commit_lock_path)
            except Exception:
                pass

    try:
        # Create our lock file
        with open(commit_lock_path, 'w') as f:
            f.write(str(os.getpid()))
            
        # Check if git is initialized
        git_check = subprocess.run(['git', 'rev-parse', '--is-inside-work-tree'], capture_output=True, text=True, cwd=WORKSPACE_DIR)
        if git_check.returncode != 0:
            return # Not a git repository
            
        # Ensure user.name and user.email are configured (common issue on fresh VMs)
        email_check = subprocess.run(['git', 'config', 'user.email'], capture_output=True, text=True, cwd=WORKSPACE_DIR)
        if not email_check.stdout.strip():
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Git email not configured. Setting a local default...")
            subprocess.run(['git', 'config', '--local', 'user.email', 'tracker-bot@stbet.local'], check=True, cwd=WORKSPACE_DIR)
            
        name_check = subprocess.run(['git', 'config', 'user.name'], capture_output=True, text=True, cwd=WORKSPACE_DIR)
        if not name_check.stdout.strip():
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Git name not configured. Setting a local default...")
            subprocess.run(['git', 'config', '--local', 'user.name', 'STBet Tracker Bot'], check=True, cwd=WORKSPACE_DIR)

        # Check if there are changes in data/
        status_check = subprocess.run(['git', 'status', '--porcelain', 'data/'], capture_output=True, text=True, cwd=WORKSPACE_DIR)
        if status_check.stdout.strip():
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Local changes detected in data/. Committing and pushing to GitHub...")
            
            # Stage changes
            subprocess.run(['git', 'add', 'data/'], check=True, cwd=WORKSPACE_DIR)
            
            # Commit changes
            commit_res = subprocess.run(['git', 'commit', '-m', 'Auto-update match records (VM source)'], capture_output=True, text=True, cwd=WORKSPACE_DIR)
            if commit_res.returncode != 0:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] Git commit failed: {commit_res.stderr.strip()}")
                return
                
            # Push changes
            push_res = subprocess.run(['git', 'push'], capture_output=True, text=True, cwd=WORKSPACE_DIR)
            if push_res.returncode == 0:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] Successfully pushed new results to GitHub Pages!")
            else:
                stderr_output = push_res.stderr.strip()
                print(f"[{datetime.now().strftime('%H:%M:%S')}] Git push failed: {stderr_output}")
                
                # Troubleshooting advice for VM
                print("\n" + "="*80)
                print("GIT PUSH TROUBLESHOOTING HINT FOR VM ENVIRONMENT:")
                print("If git push failed due to authentication, check the following options:")
                print("Option A (Token Authentication): Add your Personal Access Token (PAT) to the remote URL:")
                print("   git remote set-url origin https://<YOUR_GITHUB_TOKEN>@github.com/NorajChathuranga/STBet-Tracker.git")
                print("Option B (SSH Key Authentication): Verify that your SSH keys are set up on the VM and registered on GitHub.")
                print("   Then set remote to SSH: git remote set-url origin git@github.com:NorajChathuranga/STBet-Tracker.git")
                print("="*80 + "\n")
    except Exception as e:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Error auto-committing/pushing: {e}")
    finally:
        # Clean up our lock file
        if os.path.exists(commit_lock_path):
            try:
                os.remove(commit_lock_path)
            except Exception as e:
                print(f"Error removing commit lock file: {e}")

def main():
    load_env_file()
    migrate_root_files()
    write_initial_headers()
    
    # Check if running in GitHub Actions environment
    is_github_actions = os.environ.get('GITHUB_ACTIONS') == 'true'
    
    if is_github_actions:
        print("Running in GitHub Actions mode (single run)...")
        check_and_update()
    else:
        print("==========================================================")
        print("STBet Cricket-SuperOver Continuous Tracker (Dual Mode)")
        print(f"Data Directory: {DATA_DIR}")
        print("Checking API every 4 minutes...")
        print("Press Ctrl+C to stop.")
        print("==========================================================")
        
        # Make a check right away on startup
        if check_and_update():
            git_commit_and_push()
        
        while True:
            try:
                time.sleep(240)
                if check_and_update():
                    git_commit_and_push()
            except KeyboardInterrupt:
                print("\nTracker stopped by user.")
                break

if __name__ == '__main__':
    main()
